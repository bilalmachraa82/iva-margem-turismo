"""
Excel export module for IVA Margem cálculos.

Provides a rich OpenPyXL-powered export when the library is available.
Falls back to a simplified XLSX generator (via XlsxWriter) when OpenPyXL
is unavailable or explicitly disabled (useful in constrained environments
where the binary wheels are incompatible).
"""

import json
import locale
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import xlsxwriter

logger = logging.getLogger(__name__)


DISABLE_OPENPYXL = os.getenv("DISABLE_OPENPYXL") == "1"
OPENPYXL_AVAILABLE = False

if not DISABLE_OPENPYXL:
    try:  # pragma: no cover - simple import guard
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        try:
            from openpyxl.styles.numbers import FORMAT_CURRENCY_EUR, FORMAT_PERCENTAGE
        except ImportError:
            FORMAT_CURRENCY_EUR = '#,##0.00" €"'
            FORMAT_PERCENTAGE = '0.00%'
        OPENPYXL_AVAILABLE = True
    except Exception as exc:  # pragma: no cover - diagnostic only
        logger.warning("openpyxl indisponível (%s). Exportação avançada desativada.", exc)
else:
    logger.info("openpyxl desativado via variável DISABLE_OPENPYXL. A usar export básico.")


DISABLE_PANDAS = os.getenv("DISABLE_PANDAS") == "1"
PANDAS_AVAILABLE = False

if not DISABLE_PANDAS:
    try:  # pragma: no cover - import guard
        import pandas as pd
        PANDAS_AVAILABLE = True
    except Exception as exc:  # pragma: no cover
        logger.warning("pandas indisponível (%s). Exportação avançada desativada.", exc)
        pd = None
else:
    logger.info("pandas desativado via variável DISABLE_PANDAS. A usar export básico.")
    pd = None


class ExcelExporter:
    """Professional Excel report generator"""
    
    def __init__(self):
        self.openpyxl_enabled = OPENPYXL_AVAILABLE and PANDAS_AVAILABLE
        # Setup Portuguese locale for number formatting
        try:
            locale.setlocale(locale.LC_ALL, 'pt_PT.UTF-8')
        except locale.Error:
            try:
                locale.setlocale(locale.LC_ALL, 'Portuguese_Portugal')
            except locale.Error:
                logger.warning("Portuguese locale not available, using default")

        if self.openpyxl_enabled:
            # Define styles using OpenPyXL
            self.header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            self.header_font = Font(bold=True, color="FFFFFF", size=12)
            self.title_font = Font(bold=True, size=14)
            self.subtitle_font = Font(bold=True, size=11, color="666666")

            self.row_fill_1 = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            self.row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            self.thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            self.thick_border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        else:
            # Placeholders to avoid attribute errors
            self.header_fill = None
            self.header_font = None
            self.title_font = None
            self.subtitle_font = None
            self.row_fill_1 = None
            self.row_fill_2 = None
            self.thin_border = None
            self.thick_border = None

        # PT locale number formats (used in both modes)
        self.pt_currency_format = '#,##0.00€;[RED]-#,##0.00€'
        self.pt_number_format = '#,##0.00;[RED]-#,##0.00'
        self.pt_percentage_format = '0.00%;[RED]-0.00%'
        self.pt_date_format = 'dd/mm/yyyy'
        
    def generate(self, calculations: List[Dict], raw_data: Dict, metadata: Dict, base_dir: Optional[Path] = None) -> str:
        """
        Generate complete Excel report
        
        Args:
            calculations: List of VAT calculations
            raw_data: Original sales and costs data
            metadata: File metadata
            
        Returns:
            Path to generated Excel file
        """
        # Decide base temp directory (supports Vercel serverless using /tmp)
        if base_dir is None:
            base_dir = Path('/tmp') if os.getenv('VERCEL') else Path('temp')
        base_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = str(base_dir / f"iva_margem_{timestamp}.xlsx")
        
        try:
            if not self.openpyxl_enabled:
                generated = self._generate_basic_workbook(filename, calculations, raw_data, metadata)
                logger.info("Excel report (basic) gerado: %s", generated)
                return generated

            # Create Excel writer com OpenPyXL
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self._create_summary_sheet(writer, calculations, metadata)
                self._create_sales_sheet(writer, raw_data.get('sales', []))
                self._create_costs_sheet(writer, raw_data.get('costs', []))
                self._create_associations_sheet(writer, calculations)
                self._create_totals_sheet(writer, calculations, metadata)
                self._create_reconciliation_sheet(writer, calculations, raw_data)
                self._create_warnings_sheet(writer, calculations, raw_data)

            self._apply_formatting(filename)

            logger.info(f"Excel report generated: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}")
            raise
    
    def _build_summary_records(self, calculations: List[Dict]) -> List[Dict]:
        """Prepare summary records from calculation results."""
        summary_data = []
        for calc in calculations:
            summary_data.append({
                'Nº Documento': calc['invoice_number'],
                'Tipo': calc['invoice_type'],
                'Data': calc['date'],
                'Cliente': calc['client'],
                'Total PVP (€)': calc['sale_amount'],
                'Total Custos (€)': calc['total_allocated_costs'],
                'Margem Bruta (€)': calc['gross_margin'],
                'Taxa IVA (%)': calc['vat_rate'],
                'IVA Margem (€)': calc['vat_amount'],
                'Margem Líquida (€)': calc['net_margin'],
                'Margem (%)': calc['margin_percentage'],
                'Nº Custos': calc['cost_count']
            })
        return summary_data

    def _build_summary_dataframe(self, calculations: List[Dict]):
        if not PANDAS_AVAILABLE:
            raise RuntimeError("pandas não disponível para gerar DataFrame")
        return pd.DataFrame(self._build_summary_records(calculations))

    def _create_summary_sheet(self, writer, calculations: List[Dict], metadata: Dict):
        """Create main summary sheet with VAT calculations"""
        
        df = self._build_summary_dataframe(calculations)
        df.to_excel(writer, sheet_name='Resumo IVA Margem', index=False, startrow=3)
        
        # Get worksheet
        worksheet = writer.sheets['Resumo IVA Margem']
        
        # Add title
        worksheet['A1'] = 'Relatório IVA de Margem - Agência de Viagens'
        worksheet['A1'].font = self.title_font
        worksheet['A2'] = f"Período: {metadata.get('start_date', '')} a {metadata.get('end_date', '')}"
        worksheet['A2'].font = self.subtitle_font
        
    def _create_sales_sheet(self, writer, sales: List[Dict]):
        """Create sales detail sheet"""
        
        sales_data = []
        for sale in sales:
            sales_data.append({
                'ID': sale['id'],
                'Nº Fatura': sale['number'],
                'Data': sale['date'],
                'Cliente': sale['client'],
                'Valor s/ IVA (€)': sale.get('amount', 0),
                'IVA (€)': sale.get('vat_amount', 0),
                'Total c/ IVA (€)': sale.get('gross_total', sale.get('amount', 0)),
                'Custos Associados': len(sale.get('linked_costs', [])),
                'IDs Custos': ', '.join(sale.get('linked_costs', []))
            })
        
        df = pd.DataFrame(sales_data)
        df.to_excel(writer, sheet_name='Vendas Detalhadas', index=False, startrow=2)
        
        worksheet = writer.sheets['Vendas Detalhadas']
        worksheet['A1'] = 'Listagem de Vendas'
        worksheet['A1'].font = self.title_font
        
    def _create_costs_sheet(self, writer, costs: List[Dict]):
        """Create costs detail sheet"""
        
        costs_data = []
        for cost in costs:
            costs_data.append({
                'ID': cost['id'],
                'Fornecedor': cost['supplier'],
                'Descrição': cost.get('description', ''),
                'Data': cost['date'],
                'Nº Documento': cost.get('document_number', ''),
                'Valor s/ IVA (€)': cost.get('amount', 0),
                'IVA (€)': cost.get('vat_amount', 0),
                'Total c/ IVA (€)': cost.get('gross_total', cost.get('amount', 0)),
                'Vendas Associadas': len(cost.get('linked_sales', [])),
                'IDs Vendas': ', '.join(cost.get('linked_sales', []))
            })
        
        df = pd.DataFrame(costs_data)
        df.to_excel(writer, sheet_name='Custos Detalhados', index=False, startrow=2)
        
        worksheet = writer.sheets['Custos Detalhados']
        worksheet['A1'] = 'Listagem de Custos'
        worksheet['A1'].font = self.title_font
        
    def _create_associations_sheet(self, writer, calculations: List[Dict]):
        """Create detailed associations sheet"""
        
        associations_data = []
        for calc in calculations:
            for cost_detail in calc.get('linked_costs', []):
                associations_data.append({
                    'Fatura': calc['invoice_number'],
                    'Cliente': calc['client'],
                    'Fornecedor': cost_detail['supplier'],
                    'Descrição Custo': cost_detail['description'],
                    'Data Custo': cost_detail.get('date', ''),
                    'Custo Total (€)': cost_detail['total_amount'],
                    'Custo Alocado (€)': cost_detail['allocated_amount'],
                    'Partilhado com': f"{cost_detail['shared_with']} venda(s)",
                    'Percentagem Alocada': round((cost_detail['allocated_amount'] / cost_detail['total_amount'] * 100) if cost_detail['total_amount'] > 0 else 0, 1)
                })
        
        if associations_data:
            df = pd.DataFrame(associations_data)
            df.to_excel(writer, sheet_name='Associações Detalhadas', index=False, startrow=2)
            
            worksheet = writer.sheets['Associações Detalhadas']
            worksheet['A1'] = 'Mapa de Associações Vendas-Custos'
            worksheet['A1'].font = self.title_font
            
    def _create_totals_sheet(self, writer, calculations: List[Dict], metadata: Dict):
        """Create totals and statistics sheet"""
        
        # Calculate totals
        from .calculator import VATCalculator
        calculator = VATCalculator()
        summary = calculator.calculate_summary(calculations)
        
        # Prepare summary data
        totals_data = [
            {'': 'TOTAIS GERAIS', 'Valor': ''},
            {'': 'Total Vendas (s/ IVA)', 'Valor': f"€ {summary['total_sales']:,.2f}"},
            {'': 'Total Custos', 'Valor': f"€ {summary['total_costs']:,.2f}"},
            {'': 'Margem Bruta Total', 'Valor': f"€ {summary['total_gross_margin']:,.2f}"},
            {'': 'IVA da Margem Total', 'Valor': f"€ {summary['total_vat']:,.2f}"},
            {'': 'Margem Líquida Total', 'Valor': f"€ {summary['total_net_margin']:,.2f}"},
            {'': '', 'Valor': ''},
            {'': 'ESTATÍSTICAS', 'Valor': ''},
            {'': 'Nº Total Documentos', 'Valor': summary['documents_processed']},
            {'': 'Documentos com Lucro', 'Valor': summary['documents_with_margin']},
            {'': 'Documentos com Prejuízo', 'Valor': summary['documents_with_loss']},
            {'': 'Margem Média (%)', 'Valor': f"{summary['average_margin_percentage']}%"},
            {'': '', 'Valor': ''},
            {'': 'INFORMAÇÕES DO FICHEIRO', 'Valor': ''},
            {'': 'Empresa', 'Valor': metadata.get('company_name', '')},
            {'': 'NIF', 'Valor': metadata.get('tax_registration', '')},
            {'': 'Período', 'Valor': f"{metadata.get('start_date', '')} a {metadata.get('end_date', '')}"},
            {'': 'Moeda', 'Valor': metadata.get('currency', 'EUR')},
            {'': 'Taxa IVA Aplicada', 'Valor': f"{metadata.get('vat_rate', 23)}%"},
            {'': 'Data Cálculo', 'Valor': metadata.get('calculation_date', datetime.now().strftime('%Y-%m-%d %H:%M'))},
            {'': '', 'Valor': ''},
            {'': 'POR TIPO DE DOCUMENTO', 'Valor': ''}
        ]
        
        # Add breakdown by document type
        for doc_type, type_data in summary['by_type'].items():
            totals_data.append({
                '': f"{doc_type} ({type_data['count']} docs)",
                'Valor': f"€ {type_data['total_sales']:,.2f}"
            })
            
        df = pd.DataFrame(totals_data)
        df.to_excel(writer, sheet_name='Totais e Estatísticas', index=False, startrow=2, header=False)
        
        worksheet = writer.sheets['Totais e Estatísticas']
        worksheet['A1'] = 'Resumo e Estatísticas'
        worksheet['A1'].font = self.title_font
        
    def _apply_formatting(self, filename: str):
        """Apply professional formatting to all sheets"""

        if not self.openpyxl_enabled:
            return
        
        wb = load_workbook(filename)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers (skip sheets without headers)
            if sheet_name not in ['Totais e Estatísticas']:
                # Find header row (usually row 4 for summary, row 3 for others)
                header_row = 4 if sheet_name == 'Resumo IVA Margem' else 3
                
                # Apply header formatting
                for cell in ws[header_row]:
                    if cell.value:
                        cell.fill = self.header_fill
                        cell.font = self.header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = self.thin_border
                
                # Apply alternating row colors
                for row_num in range(header_row + 1, ws.max_row + 1):
                    fill = self.row_fill_1 if (row_num - header_row) % 2 == 1 else self.row_fill_2
                    for cell in ws[row_num]:
                        if cell.value is not None:
                            cell.fill = fill
                            cell.border = self.thin_border
                            
                            # Format values with Portuguese locale
                            header_val = str(ws.cell(header_row, cell.column).value or "").lower()

                            if isinstance(cell.value, (int, float)) and cell.column > 1:
                                if any(currency in header_val for currency in ['€', 'euro', 'valor', 'montante', 'iva']):
                                    # Portuguese currency format (1.234,56€)
                                    cell.number_format = self.pt_currency_format
                                elif any(pct in header_val for pct in ['%', 'margem', 'percentagem', 'taxa']):
                                    # Portuguese percentage format
                                    try:
                                        if float(cell.value) > 1:
                                            cell.value = float(cell.value) / 100.0
                                    except Exception:
                                        pass
                                    cell.number_format = self.pt_percentage_format
                                else:
                                    # Standard Portuguese number format (1.234,56)
                                    cell.number_format = self.pt_number_format

                                # Center align numbers
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                            elif isinstance(cell.value, str) and any(date_indicator in header_val for date_indicator in ['data', 'date']):
                                # Format date strings as Portuguese date format
                                try:
                                    from datetime import datetime
                                    # Try to parse common date formats and convert to Portuguese
                                    date_obj = None
                                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                        try:
                                            date_obj = datetime.strptime(cell.value, fmt)
                                            break
                                        except ValueError:
                                            continue

                                    if date_obj:
                                        cell.value = date_obj
                                        cell.number_format = self.pt_date_format
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                except Exception:
                                    pass  # Keep original string value
                                    
                # Add autofilter
                if ws.max_row > header_row:
                    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
                    
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                        
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Freeze panes
            if sheet_name == 'Resumo IVA Margem':
                ws.freeze_panes = 'A5'
            elif sheet_name != 'Totais e Estatísticas':
                ws.freeze_panes = 'A4'
                
        # Add footer to all sheets
        for sheet in wb.worksheets:
            sheet.oddFooter.center.text = "Powered by Accounting Advantage - &D &T"
            sheet.oddFooter.center.size = 8
            sheet.oddFooter.center.font = "Arial,Italic"
            
        # Save formatted workbook
        wb.save(filename)

    def _generate_basic_workbook(self, filename: str, calculations: List[Dict], raw_data: Dict, metadata: Dict) -> str:
        """Fallback XLSX generator when OpenPyXL/Pandas are unavailable."""

        workbook = xlsxwriter.Workbook(filename)
        try:
            summary_records = self._build_summary_records(calculations)
            if not summary_records:
                summary_records = [{
                    "Mensagem": "Sem cálculos disponíveis",
                    "Período": f"{metadata.get('start_date', '')} a {metadata.get('end_date', '')}"
                }]

            self._write_sheet_from_records(workbook, 'Resumo IVA Margem', summary_records)
            self._write_sheet_from_records(workbook, 'Vendas', raw_data.get('sales', []))
            self._write_sheet_from_records(workbook, 'Custos', raw_data.get('costs', []))
            self._write_sheet_from_records(workbook, 'Associações', calculations)
        finally:
            workbook.close()

        return filename

    def _write_sheet_from_records(self, workbook, sheet_name: str, records: List[Dict]):
        worksheet = workbook.add_worksheet(sheet_name[:31] or 'Sheet1')

        if not records:
            worksheet.write(0, 0, "Sem dados disponíveis")
            return

        headers = self._extract_headers(records)
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        for row, record in enumerate(records, start=1):
            for col, header in enumerate(headers):
                value = record.get(header)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                worksheet.write(row, col, value)

    @staticmethod
    def _extract_headers(records: List[Dict]) -> List[str]:
        headers: List[str] = []
        for record in records:
            for key in record.keys():
                if key not in headers:
                    headers.append(key)
        return headers

    def _create_reconciliation_sheet(self, writer, calculations: List[Dict], raw_data: Dict):
        """Create reconciliation sheet comparing aggregates vs per-document sums"""
        total_sales = sum(float(s.get('amount', 0) or 0) for s in raw_data.get('sales', []))
        total_costs = sum(float(c.get('amount', 0) or 0) for c in raw_data.get('costs', []))
        expected_gross = total_sales - total_costs
        allocated_sum = sum(float(c.get('total_allocated_costs', 0) or 0) for c in calculations)
        gross_sum = sum(float(c.get('gross_margin', 0) or 0) for c in calculations)

        df = pd.DataFrame([
            {'Métrica': 'Total Vendas (origem)', 'Valor': total_sales},
            {'Métrica': 'Total Custos (origem)', 'Valor': total_costs},
            {'Métrica': 'Margem Esperada (Vendas - Custos)', 'Valor': expected_gross},
            {'Métrica': 'Soma Custos Alocados (por documento)', 'Valor': allocated_sum},
            {'Métrica': 'Soma Margens (por documento)', 'Valor': gross_sum},
            {'Métrica': 'Delta Margem (Doc - Esperada)', 'Valor': gross_sum - expected_gross},
            {'Métrica': 'Delta Alocados (Doc - Custos)', 'Valor': allocated_sum - total_costs},
        ])
        df.to_excel(writer, sheet_name='Reconciliação', index=False, startrow=2)
        ws = writer.sheets['Reconciliação']
        ws['A1'] = 'Reconciliação de Totais'
        ws['A1'].font = self.title_font

    def _create_warnings_sheet(self, writer, calculations: List[Dict], raw_data: Dict):
        """Create warnings & validations sheet from calculator and raw data"""
        from .calculator import VATCalculator
        issues = VATCalculator().validate_calculations(calculations)

        sales = raw_data.get('sales', [])
        costs = raw_data.get('costs', [])
        orphan_sales = [s for s in sales if not s.get('linked_costs')]
        orphan_costs = [c for c in costs if not c.get('linked_sales')]

        rows = []
        rows.append({'Tipo': 'Info', 'Mensagem': f"Vendas sem custos: {len(orphan_sales)}"})
        rows.append({'Tipo': 'Info', 'Mensagem': f"Custos sem vendas: {len(orphan_costs)}"})
        for it in issues:
            rows.append({'Tipo': it.get('type', 'info').upper(), 'Mensagem': it.get('message', '')})

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Avisos & Validações', index=False, startrow=2)
        ws = writer.sheets['Avisos & Validações']
        ws['A1'] = 'Avisos & Validações'
        ws['A1'].font = self.title_font
