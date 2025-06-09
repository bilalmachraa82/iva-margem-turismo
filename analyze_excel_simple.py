#!/usr/bin/env python3
"""
Script simples para analisar o ficheiro Excel modelo usando openpyxl
"""
import sys
import os

def analyze_excel_simple():
    """Análise básica do ficheiro Excel"""
    
    file_path = "Modelo_IVA_Margem_Agencias_Viagens_v3.xlsm.xlsx"
    
    print("🔍 ANÁLISE DO FICHEIRO EXCEL MODELO")
    print("=" * 50)
    
    if not os.path.exists(file_path):
        print(f"❌ Ficheiro não encontrado: {file_path}")
        print("Ficheiros na pasta atual:")
        for f in os.listdir('.'):
            if f.endswith(('.xlsx', '.xlsm', '.xls')):
                print(f"  📄 {f}")
        return
    
    try:
        from openpyxl import load_workbook
        
        # Carregar o workbook
        wb = load_workbook(file_path, data_only=True)
        print(f"📋 Folhas encontradas: {wb.sheetnames}")
        print()
        
        # Analisar cada folha
        for sheet_name in wb.sheetnames:
            print(f"📊 FOLHA: {sheet_name}")
            print("-" * 30)
            
            ws = wb[sheet_name]
            
            # Dimensões da folha
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Dimensões: {max_row} linhas x {max_col} colunas")
            
            # Ler headers (primeira linha)
            headers = []
            for col in range(1, min(max_col + 1, 21)):  # Máximo 20 colunas
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Col{col}")
            
            print(f"Headers: {headers[:10]}...")  # Mostrar só os primeiros 10
            
            # Analisar dados nas primeiras linhas
            print("Primeiras 3 linhas de dados:")
            for row in range(2, min(6, max_row + 1)):  # Linhas 2-5
                row_data = []
                for col in range(1, min(len(headers) + 1, 11)):  # Primeiras 10 colunas
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is None:
                        row_data.append("")
                    elif isinstance(cell_value, (int, float)):
                        row_data.append(f"{cell_value:,.2f}" if isinstance(cell_value, float) else str(cell_value))
                    else:
                        row_data.append(str(cell_value)[:20])  # Truncar strings longas
                print(f"  Linha {row}: {row_data}")
            
            # Procurar colunas com valores monetários
            monetary_cols = []
            for col in range(1, min(max_col + 1, 21)):
                header = ws.cell(row=1, column=col).value
                if header and any(keyword in str(header).lower() for keyword in ['valor', 'montante', 'amount', 'total', 'preço', 'price', 'custo', 'venda']):
                    # Verificar se tem valores numéricos
                    has_numbers = False
                    for row in range(2, min(10, max_row + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if isinstance(cell_value, (int, float)) and cell_value != 0:
                            has_numbers = True
                            break
                    if has_numbers:
                        monetary_cols.append((col, header))
            
            if monetary_cols:
                print(f"\n💰 Colunas monetárias encontradas:")
                for col, header in monetary_cols:
                    values = []
                    for row in range(2, min(max_row + 1, 100)):  # Máximo 100 linhas
                        cell_value = ws.cell(row=row, column=col).value
                        if isinstance(cell_value, (int, float)) and cell_value != 0:
                            values.append(cell_value)
                    
                    if values:
                        total = sum(values)
                        avg = total / len(values)
                        min_val = min(values)
                        max_val = max(values)
                        
                        print(f"  {header}:")
                        print(f"    {len(values)} valores | Total: €{total:,.2f} | Média: €{avg:,.2f}")
                        print(f"    Min: €{min_val:,.2f} | Max: €{max_val:,.2f}")
                        
                        # Verificar valores negativos
                        negative_count = sum(1 for v in values if v < 0)
                        if negative_count > 0:
                            print(f"    ⚠️ {negative_count} valores negativos")
            
            print()
    
    except ImportError:
        print("❌ openpyxl não está instalado. Não foi possível analisar o Excel.")
        print("Para instalar: pip install openpyxl")
    except Exception as e:
        print(f"❌ Erro ao analisar ficheiro: {str(e)}")

if __name__ == "__main__":
    analyze_excel_simple()